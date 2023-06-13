# import xlwings as xw  # it launches excel which makes it slow down exponentially
# import xlrd  # works only for xls
import openpyxl  # this is ok
import random
import math
import struct


def generate_train_data():
    data__coordinates__x = []
    data__coordinates__y = []
    reference__x = []
    reference__y = []

    for sala in ["f8", "f10"]:
        for i in range(1, 226):
            path = f"C:\\Users\\amade\\Downloads\\SISE_2\\pomiary\\{sala.capitalize()}\\{sala}_stat_{i}.xlsx"
            print(path)
            # Specifying a sheet
            # ws = xw.Book(path).sheets['Sheet1']
            # workbook = xlrd.open_workbook(path)
            # worksheet = workbook.sheet_by_index(0)
            # Define variable to load the workbook
            workbook = openpyxl.load_workbook(path)
            # Define variable to read the active sheet:
            # worksheet = workbook.active
            worksheet = workbook["Sheet1"]

            # print(worksheet.cell_value(0, 0), end='\t')
            # Selecting data from
            # a single cell
            # data__coordinates__x = ws.range("AG2:AG127").value
            # data__coordinates__y = ws.range("AH2:AH127").value
            # reference__x = ws.range("AJ2:AJ127").value
            # reference__y = ws.range("AK2:AK127").value
            # Iterate the loop to read the cell values
            # for row_index in range(0, worksheet.max_row):
            #    for col in worksheet.iter_cols(1, worksheet.max_column):
            #        print(col[row_index].value, end="\t\t")
            #    print('')

            last_row_index = len(list(worksheet.rows))

            reference__x_symbol = "AJ"
            reference__y_symbol = "AK"
            # print(worksheet["AJ"][0].value)
            if worksheet["AJ"][0].value == "errorCode":
                reference__x_symbol = "AK"
                reference__y_symbol = "AL"

            for x in list(map(lambda c: c.value, worksheet["AG"][1:])):
                if x is None:
                    print("None detected!(x)")

            for x in list(map(lambda c: c.value, worksheet[reference__x_symbol][1:])):
                if x is None:
                    print("None detected! (reference_x)")

            data__coordinates__x += list(map(lambda c: c.value, worksheet["AG"][1:]))
            data__coordinates__y += list(map(lambda c: c.value, worksheet["AH"][1:]))
            reference__x += list(map(lambda c: c.value, worksheet[reference__x_symbol][1:]))
            reference__y += list(map(lambda c: c.value, worksheet[reference__y_symbol][1:]))

    not_none_indices = list(filter(lambda ix: data__coordinates__x[ix] is not None, range(len(data__coordinates__x))))

    # print(not_none_indices)
    # print(f"Len of data__coordinates__x: {len(data__coordinates__x)}")
    # print(f"Len of data__coordinates__y: {len(data__coordinates__y)}")
    # print(f"Len of reference__x: {len(reference__x)}")
    # print(f"Len of reference__y: {len(reference__y)}")
    # print(f"Len of not_none_indices: {len(not_none_indices)}")

    data__coordinates__y = list(map(lambda ix: data__coordinates__y[ix], not_none_indices))
    reference__x = list(map(lambda ix: reference__x[ix], not_none_indices))
    reference__y = list(map(lambda ix: reference__y[ix], not_none_indices))
    data__coordinates__x = list(filter(lambda x_: x_ is not None, data__coordinates__x))

    # for ind in range(len(reference__y)):
    #     if reference__y[ind] is None:
    #         print(f"reference__y NONE DETECTED AT ind={ind}")

    print(f"Len of data__coordinates__x: {len(data__coordinates__x)}")
    print(f"Len of data__coordinates__y: {len(data__coordinates__y)}")
    print(f"Len of reference__x: {len(reference__x)}")
    print(f"Len of reference__y: {len(reference__y)}")

    # print(data__coordinates__x)

    # for i in range(len(data__coordinates__y)):
    #    if data__coordinates__y[i] is None:
    #        print(f"None at index {i}")

    # print(data__coordinates__y)
    # print(reference__x)
    # print(reference__y)

    indices = list(range(len(data__coordinates__x)))
    # random.shuffle(indices) ###
    # print(indices)

    data__coordinates__x = list(map(lambda ix: data__coordinates__x[ix], indices))
    data__coordinates__y = list(map(lambda ix: data__coordinates__y[ix], indices))
    reference__x = list(map(lambda ix: reference__x[ix], indices))
    reference__y = list(map(lambda ix: reference__y[ix], indices))

    inputs = [0] * 2 * len(data__coordinates__x)
    outputs = [0] * 2 * len(reference__x)
    inputs[0::2] = data__coordinates__x
    inputs[1::2] = data__coordinates__y
    outputs[0::2] = reference__x
    outputs[1::2] = reference__y

    for x in data__coordinates__x:
        if x is None:
            print("X NONE DETECTED")

    for y in data__coordinates__y:
        if y is None:
            print("y NONE DETECTED")

    for x_ref in reference__x:
        if x_ref is None:
            print("X_REF NONE DETECTED")

    for y_ref in reference__x:
        if y_ref is None:
            print("Y_REF NONE DETECTED")

    # print(inputs[0])
    # print(inputs[1])
    # print(inputs[2])
    # print(inputs[3])

    # print(data__coordinates__x[0:5])
    # print(data__coordinates__y[0:5])
    # print(reference__x[0:5])
    # print(reference__y[0:5])

    # print(data__coordinates__x[0])
    # print(data__coordinates__y[0])
    # print(data__coordinates__x[1])
    # print(data__coordinates__y[1])

    print(inputs[0:5])
    print(outputs[0:5])

    inputs = list(map(lambda val: -1.0 + 2.0 / (1.0 + math.exp(-0.001 * val)), inputs))
    outputs = list(map(lambda val: -1.0 + 2.0 / (1.0 + math.exp(-0.001 * val)), outputs))

    print(inputs[0:5])
    print(outputs[0:5])

    # inputsFile = open("inputs.txt", "w+")
    # inputsFile.write(str(inputs))
    # inputsFile.close()

    # outputsFile = open("outputs.txt", "w+")
    # outputsFile.write(str(outputs))
    # outputsFile.close()

    with open('inputs.bin', 'wb') as file:
        # Convert the float array to binary representation
        binary_data = struct.pack('d' * len(inputs), *inputs)

        # Write the binary data to the file
        file.write(binary_data)

    with open('outputs.bin', 'wb') as file:
        # Convert the float array to binary representation
        binary_data = struct.pack('d' * len(outputs), *outputs)

        # Write the binary data to the file
        file.write(binary_data)

    print(f"Training samples count: {0.5 * len(inputs)}")
    print(f"Training samples count: {0.5 * len(outputs)}")

    print(inputs[0])
    print(inputs[1])
    print(inputs[2])
    print(inputs[3])